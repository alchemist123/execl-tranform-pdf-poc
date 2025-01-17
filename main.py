from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import math
import os
from datetime import datetime

def create_output_directory():
    os.makedirs('out-files', exist_ok=True)

def get_output_filepath():

    current_date = datetime.now().strftime("%Y-%m-%d")
    return os.path.join('out-files', f'{current_date}.pdf')

def load_excel_file(filepath):

    wb = load_workbook(filepath)
    return wb.active

def setup_pdf_document(output_path):

    return SimpleDocTemplate(output_path, pagesize=landscape(letter))

def calculate_pagination(max_col, max_row, doc_width):

    tuner_columns_number = 8
    col_width = doc_width / tuner_columns_number
    cols_per_page = min(tuner_columns_number, max_col)
    num_batches = math.ceil(max_col / cols_per_page)
    
    rows_per_page = 30
    num_row_batches = math.ceil(max_row / rows_per_page)
    
    return cols_per_page, num_batches, rows_per_page, num_row_batches

def get_batch_bounds(batch_idx, items_per_page, max_items):

    start = batch_idx * items_per_page + 1
    end = min((batch_idx + 1) * items_per_page, max_items)
    return start, end

def extract_batch_data(ws, start_row, end_row, start_col, end_col):

    data = []
    
    # Add header row
    header_row = [ws.cell(row=1, column=col).value for col in range(start_col, end_col + 1)]
    data.append(header_row)
    
    # Add data rows
    for row in range(max(2, start_row), end_row + 1):
        data_row = [ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]
        data.append(data_row)
    
    return data

def create_styled_table(data):

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    return table

def create_page_number(page_counter, col_batch):

    page_num = str(page_counter) if col_batch == 0 else f"{page_counter}.{col_batch}"
    return Paragraph(f"Page {page_num}", getSampleStyleSheet()['Normal'])

def excel_to_pdf():
    try:
        create_output_directory()
        output_pdf_path = get_output_filepath()
        
        # Load Excel file with input excel file path
        ws = load_excel_file("input.xlsx")
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Setup PDF document with landscape orientation
        doc = setup_pdf_document(output_pdf_path)
        elements = []
        
        # Calculate pagination parameters for current excel file
        cols_per_page, num_batches, rows_per_page, num_row_batches = calculate_pagination(
            max_col, max_row, doc.width
        )
        
        page_counter = 1
        
        for row_batch in range(num_row_batches):
            start_row, end_row = get_batch_bounds(row_batch, rows_per_page, max_row)
            
            for col_batch in range(num_batches):
                start_col, end_col = get_batch_bounds(col_batch, cols_per_page, max_col)
                
                # Extract and format batch data for current batch
                data = extract_batch_data(ws, start_row, end_row, start_col, end_col)
                table = create_styled_table(data)
                
                # Add elements to PDF document
                elements.append(table)
                elements.append(Spacer(1, 20))
                elements.append(create_page_number(page_counter, col_batch))
                elements.append(PageBreak())
            
            page_counter += 1
        
        # Build PDF document
        doc.build(elements)
        return True
        
    except Exception as e:
        print(f"Error converting Excel to PDF: {str(e)}")
        return False

if __name__ == "__main__":
    excel_to_pdf()